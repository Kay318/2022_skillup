from csv import excel
from openpyxl.styles import Alignment
from re import I
from typing import List
import openpyxl as xl
import xlwings as wings
import pandas
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet
from Database.DB import DBManager
from Settings import Setup as sp

class excelModul(DBManager):
    
    def __init__(self, save_path, lang_List, new_set_difference) -> excel:
        super().__init__()
        
        self.set_row = 1
        self.start_column = 1
        self.lang_cnt = 1
        self.sp = sp.Settings()
        
        if (new_set_difference):

            wb = xl.Workbook()

            for lang in lang_List:

                active = wb.create_sheet(title=lang)

                # 초기세팅
                active.cell(row = self.set_row, # 현재 진행상황
                            column= self.lang_cnt, # 평가한 나라 개수 영향
                            value=f'{lang}_이미지')

                idx = self.start_column + self.lang_cnt # 언어 1개 고정
                for val in self.evaluation_len(key="Test_List"):
                    active.cell(row = self.set_row, # 현재 진행상황
                                column= idx, # 평가한 나라 개수 영향
                                value=val)
                    idx = idx + 1

                jdx = idx
                for val in self.evaluation_len(key="Field"):
                    active.cell(row = self.set_row, # 현재 진행상황
                                column= jdx, # 평가한 나라 개수 영향
                                value=val)
                    jdx = jdx + 1

                active.cell(row = self.set_row, # 현재 진행상황
                            column= jdx, # 평가한 나라 개수 영향
                            value="버전정보")

                self.set_row = self.set_row + 1
                self.excel_data_input(active = active, lang = lang)


            for val_row in active.rows:
                for cell in val_row:

                    cell.alignment = Alignment(horizontal="center", vertical="center")

            wb.save(f'{save_path}\\excelTest.xlsx')

        else:
            # 기존 엑셀 생성

            wb = wings.Book(save_path)
            for lang in lang_List:
                sheet = wb.sheets[lang]
                for lang in sheet:
                    print(lang)
                

    def 이미지개수당셀생성(self, lang) -> List:
        """
        1. DB 받아오기
        2. 저장된 DB중 이미지 경로 있는것만 구별시키기
        3. 이미지 개수를 확인하고 : return 해당경로를 순서대로 List 반환
        """

        self.c.execute(f'SELECT * FROM {lang}')
        dataList = self.c.fetchall()
        print(f"ex : {dataList}")
        idx = 0 # "평가결과저장된데이터중 경로위치"
        img_pathList = []

        # 행마다 차래로
        for data in dataList:
            
            print(f"이미지 : {data[0]}")
            if data != "":
                # clume서치
                img_pathList.append(data[idx])

        return img_pathList

    def cell_search_val(self, sequence, row, lang) -> List:
        """
        1. DB 받아오기
        2. 저장된 DB중 이미지 경로 있는것만 구별시키기
        3. 이미지 개수를 확인하고 : return 해당경로를 순서대로 List 반환
        """

        self.c.execute(f'SELECT * FROM {lang}')
        dataList = self.c.fetchall()
        print(f'{dataList[sequence]} : {dataList[sequence][row]}')

        return dataList[sequence][row]

    def evaluation_len(self, key) -> List:

        result_val, result_val2 = self.sp.read_ini__test(table = key)
        
        result_cellList = []

        for i in result_val:
            result_cellList.append(i)
  
        return result_cellList

    def excel_data_input(self, active : Worksheet, lang):
        # 데이터 추가
        
        pre = 0
        gap_count = 0
        column_interval = 1
        active.column_dimensions['A'].width = 40

        for path in self.이미지개수당셀생성(lang = lang):

            if (path != ""):

                img = Image(path)
                img.width = 320
                img.height = 220
                
                active.add_image(img=img, anchor=f"A{self.set_row}") # 이미지 추가
                active.row_dimensions[self.set_row].height = 185
                
                # 평가 결과
                gap_count = column_interval
                for idx in range(self.lang_cnt, len(self.evaluation_len(key="Test_List")) + self.lang_cnt):

                    print(f'idx : {idx}')
                    if (idx <= len(self.evaluation_len(key="Test_List")) + self.lang_cnt):

                        active.cell(row = self.set_row,
                            column= idx + gap_count, # index 2
                            value = self.cell_search_val(sequence = pre, row = idx, lang = lang))

                # 라벨 평가
                gap_count = gap_count + len(self.evaluation_len(key="Test_List"))
                for idx in range(self.lang_cnt, len(self.evaluation_len(key="Field")) + self.lang_cnt):

                    print(f'idx : {idx}')
                    if (idx <= len(self.evaluation_len(key="Field")) + self.lang_cnt):

                        print(f'value : {self.cell_search_val(sequence = pre, row = idx + len(self.evaluation_len(key="Test_List")), lang = lang)}')
                        active.cell(row = self.set_row,
                            column= idx + gap_count, # index 2
                            value = self.cell_search_val(sequence = pre, row = idx + len(self.evaluation_len(key="Test_List")), lang = lang))
            
                # 버전 정보
                gap_count = gap_count + len(self.evaluation_len(key="Field"))

                print(f'버전정보 : {self.cell_search_val(sequence = pre, row = gap_count, lang = lang)}')

                active.cell(row = self.set_row,
                            column= idx + gap_count - 2, # index 2
                            value = self.cell_search_val(
                                sequence = pre, 
                                row = gap_count, 
                                lang = lang))

            self.set_row = self.set_row + 1
            pre = pre + 1