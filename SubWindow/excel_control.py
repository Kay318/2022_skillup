from csv import excel
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import borders
from openpyxl.styles.fonts import Font
from typing import List
import openpyxl as xl
import win32com.client as win32
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet
from DataBase.DB import DBManager
from PIL import Image as IMG
from Settings import Setup as sp
import os
import string
import time
from Log import LogManager
from progressBar import ProgressApp
from PyQt5.QtCore import QThread, QRunnable, QThreadPool, pyqtSlot, QObject, pyqtSignal, QEventLoop, QTimer
from PyQt5.QtWidgets import QApplication, QMessageBox, QWidget

class excelRun(QWidget ,DBManager):

    def __init__(self, save_path, lang_List, new_set_difference) -> None:
        super().__init__()

        wb = object
        COUNT = 0
        ADD_COUNT = 0
        self.save_path = save_path
        print(save_path)
        for lang in lang_List:
            COUNT = COUNT + len(self.imgCellCount(lang= lang))

        if new_set_difference:
            wb = xl.Workbook()
            ADD_COUNT = 50
        else:
            ADD_COUNT = 70
            COUNT = COUNT + ADD_COUNT

        if COUNT > 100:

            for i in range(COUNT // 100):
                COUNT = COUNT + ADD_COUNT
        
        if (self.setting_Verification(langList = lang_List)):
            self.progress_Thread = QThread()
            self.progress_Thread.start()
            self.worker = ProgressApp(time=int(COUNT), new_set_difference = new_set_difference, save_path= self.save_path, wb= wb)
            self.worker.moveToThread(self.progress_Thread)

            self.exModul_Thread = QThread()
            self.exModul_Thread.start()
            self.exModuls = excelModul(save_path = self.save_path, lang_List = lang_List, new_set_difference = new_set_difference, wb= wb)
            self.exModuls.moveToThread(self.exModul_Thread)

    def setting_Verification(self, langList):
        path = str(os.path.dirname(self.save_path))

        result = False
        
        if os.path.isdir(path):
            result = True
        else:
            btnReply = QMessageBox.warning(self, "주의", f"{path} 경로가 존재하지 않습니다.", QMessageBox.Cancel, QMessageBox.Cancel)
            LogManager.HLOG.info("언어 설정 팝업에서 존재하지 않는 경로 알림 표시")
            
            if btnReply == QMessageBox.Cancel:
                result = False
                return result

        if result:

            for lang in langList:
                
                path = os.path.dirname(self.imgCellCount(lang)[0])
                if os.path.isdir(path):
                    result = True
                else:
                    btnReply = QMessageBox.warning(self, "주의", f"{self.imgCellCount(lang)[0]} 경로가 존재하지 않습니다.", QMessageBox.Cancel, QMessageBox.Cancel)
                    LogManager.HLOG.info("언어 설정 팝업에서 존재하지 않는 경로 알림 표시")
                    if btnReply == QMessageBox.Cancel:
                        result = False
                        return result
    
        return result

    def imgCellCount(self, lang) -> List:
        """
        1. DB 받아오기
        2. 저장된 DB중 이미지 경로 있는것만 구별시키기
        3. 이미지 개수를 확인하고 : return 해당경로를 순서대로 List 반환
        """

        self.c.execute(f"SELECT * FROM '{lang}'")
        dataList = self.c.fetchall()
    
        idx = 0 # "평가결과저장된데이터중 경로위치"
        img_pathList = []

        # 행마다 차래로
        for data in dataList:
            
            if data != "":
                # clume서치
                img_pathList.append(data[idx])

        return img_pathList

class excelModul(QObject, DBManager):
    
    def __init__(self, save_path, lang_List, new_set_difference, wb:object) -> excel:
        super().__init__()

        loop = QEventLoop()
        QTimer.singleShot(1000, loop.quit)

        # 초기설정
        self.IMG_WIDTHSIZE = None # 이미지 가로 크기
        self.IMG_HEIGHTSIZE = None # 이미지 세로 크기
        self.IMG_SHEET_HEIGHTSIZE = None # 이미지 셀 행 크기
        self.SHEET_WIDTHSIZE = None # 시트 셀 기본 열 크기
        self.SHEET_HEIGHTSIZE = None # 시트 셀 기본 행 크기
        self.SHEET_WIDTH_SHORTSIZE = 15 # 시트 열 기본 작은 크기
        self.TABLE_CELL_COLOR = 43 # 테이블 컬러
        
        self.set_row = 1
        self.start_column = 1
        self.lang_cnt = 1
        self.sp = sp.Settings()
        self.get_lang_List = lang_List
        self.wb = wb
        
        self.threadPool = QThreadPool().globalInstance()

        self.column = []

        self.excel_setup()
        print(f'self.IMG_WIDTHSIZE : {self.IMG_WIDTHSIZE}')
        print(f'self.IMG_HEIGHTSIZE : {self.IMG_HEIGHTSIZE}')
        print(f'self.IMG_SHEET_HEIGHTSIZE : {self.IMG_SHEET_HEIGHTSIZE}')
        print(f'self.SHEET_WIDTHSIZE : {self.SHEET_WIDTHSIZE}')
        print(f'self.SHEET_HEIGHTSIZE : {self.SHEET_HEIGHTSIZE}')

        for i in range(65, 91):
            self.column.append(chr(i))

        if (new_set_difference):

            ws = self.wb.active
            ws.title = "RESULT"
            self.history_rows = 1

            for lang in lang_List:

                self.create_sheet_history(lang=lang, ws= ws)
    
                active = self.wb.create_sheet(title=lang)

                self.excel_data_input(active = active, lang = lang)
                
                self.set_cellStyle(active= active, idx= 1, fix=None)
                
            # print(f'save 저장 전')
            # wb.save(f'{save_path}\\excelTest.xlsx')
            # print(f'save 저장 후')

        else:

            TITLE_TERGET = 1
            self.historyUpdate_rows = 2

            excel = win32.Dispatch("Excel.Application")
            excel.Visible = False
            self.wb = excel.Workbooks.Open(save_path)

            for lang in lang_List:
                self.update_sheet_history(lang, self.wb.Worksheets("RESULT"), TITLE_TERGET)
                ws = self.wb.Worksheets(lang)

                kr_list = [] # 현재 엑셀 데이터
                idx = 1

                while(ws.Cells(TITLE_TERGET, idx).Value != None):
                    value = "" if "" == ws.Cells(TITLE_TERGET, idx).Value else ws.Cells(TITLE_TERGET, idx).Value
                    
                    kr_list.append(value)
                    idx = idx + 1

                self.__equalsVerification__(kr_list=kr_list, ws = self.wb.Worksheets(lang), lang = lang, terget=TITLE_TERGET)
                QApplication.processEvents()
                
            excel.Quit()

        loop.exec_()
        print(f'루프종료')
        loop.exit()

    def excel_setup(self):
        excel_setList, _ = self.sp.read_setup(table = "Excel_Setting")

        self.IMG_WIDTHSIZE = int(excel_setList[0]) # 이미지 가로 크기
        self.IMG_HEIGHTSIZE = int(excel_setList[1]) # 이미지 세로 크기
        self.IMG_SHEET_HEIGHTSIZE = int(excel_setList[2]) # 이미지 셀 행 크기
        self.SHEET_WIDTHSIZE = int(excel_setList[3]) # 시트 셀 기본 열 크기
        self.SHEET_HEIGHTSIZE = int(excel_setList[4]) # 시트 셀 기본 행 크기
        

    def create_imgCellCount(self, lang) -> List:
        """
        1. DB 받아오기
        2. 저장된 DB중 이미지 경로 있는것만 구별시키기
        3. 이미지 개수를 확인하고 : return 해당경로를 순서대로 List 반환
        """

        self.c.execute(f'SELECT * FROM {lang}')
        dataList = self.c.fetchall()
    
        idx = 0 # "평가결과저장된데이터중 경로위치"
        img_pathList = []

        # 행마다 차래로
        for data in dataList:
            
            if data != "":
                # clume서치
                img_pathList.append(data[idx])

        return img_pathList

    def cell_search_val(self, sequence, columns, lang) -> List:
        """
        1. DB 받아오기
        2. 저장된 DB중 이미지 경로 있는것만 구별시키기
        3. 이미지 개수를 확인하고 : return 해당경로를 순서대로 List 반환
        """

        self.c.execute(f'SELECT * FROM {lang}')
        dataList = self.c.fetchall()
        print(f'{dataList[sequence]} : {dataList[sequence][columns]}')

        return dataList[sequence][columns]

    def overloading_cell_search_val(self, sequence, lang) -> List:
        """
        1. DB 받아오기
        2. 저장된 DB중 이미지 경로 있는것만 구별시키기
        3. 이미지 개수를 확인하고 : return 해당경로를 순서대로 List 반환
        """

        self.c.execute(f'SELECT * FROM {lang}')
        dataList = self.c.fetchall()

        return dataList[sequence]

    def evaluation_len(self, key) -> List:

        result_val, result_val2 = self.sp.read_setup(table = key)
        
        result_cellList = []

        for i in result_val:
            result_cellList.append(i)
  
        return result_cellList

    def excel_data_input(self, active : Worksheet, lang):
        # 데이터 추가
        
        lang_books = []

        cell_idx = 1
        cell_rows = 1
        for val in self.create_tupleBooks(lang=lang):
            lang_books.append(val)

        # 태그적용
        for val in range(0, len(lang_books)):
            active.cell(row = cell_rows, # 현재 진행상황
                        column= cell_idx, # 평가한 나라 개수 영향
                        value=lang_books[val])
            cell_idx = cell_idx + 1

        self.set_cellStyle(active= active, idx = cell_rows, fix= None)

        # 언어 이미지명 testList Field 버전
        
        sequence = 0
        cell_rows = cell_rows + 1
        for path in self.create_imgCellCount(lang = lang):

            if (path != ""):

                columns= 0 + 1
                columnsVal = 0
                
                try:
                    img = Image(path)
                    img.width = self.IMG_WIDTHSIZE
                    img.height = self.IMG_HEIGHTSIZE
                    
                    active.add_image(img=img, anchor=f"A{cell_rows}") # 이미지 추가
                    active.row_dimensions[cell_rows].height = self.IMG_SHEET_HEIGHTSIZE
                    active.cell(row = cell_rows,
                                column= columns,
                                value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang))
                except:
                    active.row_dimensions[cell_rows].height = self.IMG_SHEET_HEIGHTSIZE
                    active.cell(row = cell_rows,
                                column= idx + 1,
                                value = "경로 없음")
                active.cell(row = cell_rows,
                        column= columns,
                        value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang))
                
                # 평가 결과
                for idx in range(0, len(self.evaluation_len(key="Test_List"))):

                    columns= columns + 1
                    columnsVal = columnsVal + 1
                    print(f'columns : {columns}')
                    print(f'Value : {self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)}')
                    value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)

                    active.cell(row = cell_rows,
                        column= columns,
                        value = ("" if "" == value else value))

                # 라벨 평가
                for idx in range(0, len(self.evaluation_len(key="Field"))):

                    columns= columns + 1
                    columnsVal = columnsVal + 1
                    print(f'columns : {columns}')
                    print(f'Value : {self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)}')
                    value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)

                    active.cell(row = cell_rows,
                        column= columns,
                        value = ("" if "" == value else value))

            
                # 버전 정보
                print(f'columns : {columns + 1}')
                print(f'Value : {self.cell_search_val(sequence = sequence, columns = columnsVal + 1, lang = lang)}')
                value = self.cell_search_val(sequence = sequence, columns = columnsVal + 1, lang = lang)
                active.cell(row = cell_rows,
                            column= columns + 1,
                            value = "" if "" == value else value)

                cell_rows = cell_rows + 1
            sequence = sequence + 1

        cell_rows = cell_rows + 1


    def set_cellStyle(self, active : Worksheet, idx : int, fix:str):

        i = 0
        for column in active.columns:
            
            if fix == "RESULT":
                if (i == 1 or i > len(self.evaluation_len(key="Test_List")) + 1):
                    active.column_dimensions[self.column[i]].width = self.SHEET_WIDTHSIZE
                else:
                    active.column_dimensions[self.column[i]].width = self.SHEET_WIDTH_SHORTSIZE
            else:
                if (i == 0 or i > len(self.evaluation_len(key="Test_List"))):
                    active.column_dimensions[self.column[i]].width = self.SHEET_WIDTHSIZE
                else:
                    active.column_dimensions[self.column[i]].width = self.SHEET_WIDTH_SHORTSIZE

            active[f'{self.column[i]}{idx}'].font = Font(size=11)

            i = i + 1

        active.row_dimensions[self.set_row].height = self.SHEET_HEIGHTSIZE

        upper = [f'{i}{idx}' for i in string.ascii_uppercase]
        
        for val_row in active.rows:
            for cell in val_row:

                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                str_cell = str(cell)
                str_cell = str_cell[str_cell.find(".") + 1:str_cell.find(">")]
                print(f'cell : {str_cell}')
                print(f'upper : {upper}')
                if (str_cell in upper):
                    cell.fill = PatternFill(patternType="solid", fgColor="99CC00")

                cell.border = Border(
                            right=Side(border_style=borders.BORDER_THIN,
                                    color='000000'),
                            bottom=Side(border_style=borders.BORDER_THIN,
                                        color='000000'))
                time.sleep(0.005)
                QApplication.processEvents()


    def set_Win32com_cellStyle(self, ws : object, terget : int, cell_idx : int, heightSize : int):
        if heightSize != None:

            if cell_idx > 1:
                ws.Rows(cell_idx).RowHeight = heightSize  # 선택 영역 행 크기 설정
        else:
            ws.Rows(cell_idx).RowHeight = self.SHEET_HEIGHTSIZE  # 선택 영역 행 크기 설정

        rng = ws.UsedRange # 사용 영역 선택
    
        #테두리 설정
        rng.Borders.LineStyle = 1 #선 스타일
        rng.Borders.ColorIndex = 1 #선 색상 : Black
        rng.Borders.Weight = 2 #선 굵기
        ws.Cells(terget, cell_idx).Font.Size = 11
        ws.Cells(terget, cell_idx).VerticalAlignment = -4108  #가운데 정렬(수직)
        ws.Cells(terget, cell_idx).HorizontalAlignment = -4108  #가운데 정렬(수평)
    
    def create_tupleBooks(self, lang) -> list:
        """
        언어 마다 엑셀 타이틀 항목 반환 함수
        * ini에 영향 끼침.

        return : Title list result
        """

        tupleBooks = {}

        for key in self.get_lang_List:
            
            all_List = []
            all_List.append(f'{key}_이미지')

            for val in self.evaluation_len(key="Test_List"):
                all_List.append(val)

            for val in self.evaluation_len(key="Field"):
                all_List.append(val)
            
            all_List.append("버전정보")

            tupleBooks.setdefault(key, all_List)
        
        return tupleBooks.get(lang)

    def __equalsVerification__(self, kr_list : list, ws : object, lang : str, terget : int):
        """
        구별 : 차별필요
        """
        lang_books = self.create_tupleBooks(lang=lang)

        if len(kr_list) > len(lang_books):

            del_ColCount = len(kr_list) - (len(kr_list) - len(lang_books)) + 1
            for i in range(0, len(kr_list) - len(lang_books)):
                del kr_list[len(kr_list) - 1]
                ws.Columns(del_ColCount).EntireColumn.Delete()
        elif len(kr_list) < len(lang_books):
            for i in range(0, len(lang_books) - len(kr_list)):
                kr_list.append("NULL")

        # 태그적용
        for val in range(0, len(kr_list)):
            cell_idx = val + 1
            if kr_list[val] != lang_books[val]:
                ws.Cells(terget, cell_idx).Value = lang_books[val]
                kr_list[val] = lang_books[val]
            else:
                ws.Cells(terget, cell_idx).Value = kr_list[val]
            ws.Cells(terget, cell_idx).Interior.ColorIndex = self.TABLE_CELL_COLOR # 색상적용
            self.set_Win32com_cellStyle(ws = ws, terget = terget, cell_idx = cell_idx, heightSize = None)
            print(f'time : {val}')

        # 엑셀 데이터 적용
        for key_terget in range(0, len(self.create_imgCellCount(lang=lang))):
            
            cellCount = key_terget + 2
            if cellCount <= len(self.create_imgCellCount(lang=lang)):
                try:
                    path = self.create_imgCellCount(lang=lang)[key_terget]
                    img = IMG(key_terget)
                    img.width = self.IMG_WIDTHSIZE
                    img.height = self.IMG_HEIGHTSIZE
                    
                    image = ws.Shapes.AddPicture(path, False, True, cellCount, 1, img.width, img.height)
                    ws.Cells(cellCount, 1).Value = path
                except:
                    ws.Cells(cellCount, 1).Value = "경로 없음"
                time.sleep(0.005)

                data_List = list(self.overloading_cell_search_val(sequence= key_terget, lang= lang))

                for val in range(0, len(kr_list)):
                    if kr_list[val] != data_List[val]:
                        cell_idx = val + 1
                        ws.Cells(cellCount, cell_idx).Value = ("" if "" == data_List[val] else data_List[val])
                        self.set_Win32com_cellStyle(ws = ws, terget = cellCount, cell_idx = cell_idx, heightSize = self.IMG_SHEET_HEIGHTSIZE)
                print(f'진행도 : {cellCount}/{len(self.create_imgCellCount(lang=lang))}')
                time.sleep(0.005)
                QApplication.processEvents()
            
    def create_sheet_history(self, lang:str, ws:object):
        """
        연속적으로 생성 할 수 있는 규칙 필요
        """

        lang_books = []
        lang_books.append("언어")

        
        cell_idx = 1
        cell_row = self.history_rows
        for val in self.create_tupleBooks(lang=lang):
            lang_books.append(val)

        # 태그적용
        for val in range(0, len(lang_books)):
            ws.cell(row = self.history_rows, # 현재 진행상황
                        column= cell_idx, # 평가한 나라 개수 영향
                        value=lang_books[val])
            cell_idx = cell_idx + 1

        # 언어 이미지명 testList Field 버전
        
        sequence = 0
        self.history_rows = self.history_rows + 1
        for path in self.create_imgCellCount(lang = lang):

            if (path != ""):
                
                print(f'terget : {self.overloading_cell_search_val(sequence = sequence, lang= lang)}')
                if "FAIL" in self.overloading_cell_search_val(sequence = sequence, lang= lang) or\
                    "N/A" in self.overloading_cell_search_val(sequence = sequence, lang= lang) or\
                        "N/T" in self.overloading_cell_search_val(sequence = sequence, lang= lang):

                    columns= 0 + 1
                    print(f'columns : {columns}')
                    print(f'Value : {lang}')
                    ws.cell(row = self.history_rows,
                            column= columns,
                            value = lang)

                    columns = columns + 1
                    columnsVal = 0
                    print(f'columns : {columns}')
                    print(f'Value : {self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)}')
                    ws.cell(row = self.history_rows,
                            column= columns,
                            value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang))
                    
                    # 평가 결과
                    for idx in range(0, len(self.evaluation_len(key="Test_List"))):

                        columns= columns + 1
                        columnsVal = columnsVal + 1
                        print(f'columns : {columns}')
                        print(f'Value : {self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)}')
                        value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)

                        ws.cell(row = self.history_rows,
                            column= columns,
                            value = ("" if "" == value else value))

                    # 라벨 평가
                    for idx in range(0, len(self.evaluation_len(key="Field"))):

                        columns= columns + 1
                        columnsVal = columnsVal + 1
                        print(f'columns : {columns}')
                        print(f'Value : {self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)}')
                        value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)

                        ws.cell(row = self.history_rows,
                            column= columns,
                            value = ("" if "" == value else value))

                
                    # 버전 정보
                    print(f'columns : {columns + 1}')
                    print(f'Value : {self.cell_search_val(sequence = sequence, columns = columnsVal + 1, lang = lang)}')
                    value = self.cell_search_val(sequence = sequence, columns = columnsVal + 1, lang = lang)
                    ws.cell(row = self.history_rows,
                                column= columns + 1,
                                value = "" if "" == value else value)

                    self.history_rows = self.history_rows + 1
                sequence = sequence + 1

        self.set_cellStyle(active= ws, idx = cell_row, fix= "RESULT")
        self.history_rows = self.history_rows + 1

    def update_sheet_history(self, lang:str, ws:object, integer:int):

        row = self.historyUpdate_rows
        colCount = 1
        colList = []
        # 새로운 데이터 추가 부여필요
        tuples = []
        tuples.append("언어")
        for tuple in self.create_tupleBooks(lang= lang):
            tuples.append(tuple)

        self.set_Win32com_cellStyle(ws = ws, terget = row - 1, cell_idx = colCount, heightSize = None)
        while(ws.Cells(row, integer).Value == lang):
            print(f"확인 : {ws.Cells(row, integer).Value}")
            row = row + 1

        while(ws.Cells(self.historyUpdate_rows - 1, colCount).Value != None):
            print(f'colCountValue : {ws.Cells(integer, colCount).Value}')
            colList.append(ws.Cells(self.historyUpdate_rows - 1, colCount).Value)
            colCount = colCount + 1
        colCount = colCount - 1

        print(f'colCount : {colCount}')
        print(f'lang : {lang}')
        print(f'DB_colCount : {len(self.create_tupleBooks(lang= lang))}')
        # 컬럼개수의 변동이 있었는가?
        if (colList == tuples):
            getPathList = []
            pathList = []
            
            for row_idx in range(self.historyUpdate_rows, row):
                print(f"ws.Cells(row_idx, integer).Value : {ws.Cells(row_idx, integer).Value}")
                if (ws.Cells(row_idx, integer).Value == lang):
                    getPathList.append(os.path.basename(ws.Cells(row_idx, 2).Value)) # 파일명만 추출

            for idx in range(0, len(self.create_imgCellCount(lang = lang))):

                data_List = list(self.overloading_cell_search_val(sequence= idx, lang= lang))
                if ("FAIL" in data_List or "N/A" in data_List or "N/T" in data_List):
                    pathList.append(os.path.basename(self.create_imgCellCount(lang)[idx])) # 파일명만 추출모음
                
            print(f'getPathList : {len(getPathList)}')
            print(f'pathList : {len(pathList)}')

            if len(getPathList) != len(pathList):

                print(f'pathList : {len(pathList)}')
                print(f'getPathList : {len(getPathList)}')

                char_lenght = len(pathList) - len(getPathList)
                print(f'char_lenght : {char_lenght}')
                upper = [f'{i}' for i in string.ascii_uppercase]
                if (char_lenght > 0):
                    
                    new_style_terget = f"{upper[integer - 1]}{self.historyUpdate_rows}:{upper[len(self.create_tupleBooks(lang = lang))]}{self.historyUpdate_rows}"
                    for i in range(0, char_lenght):
                        ws.Rows(self.historyUpdate_rows).Insert()
                        ws.Rows(self.historyUpdate_rows).RowHeight = self.IMG_SHEET_HEIGHTSIZE
                        ws.Range(f"{new_style_terget}").Interior.ColorIndex = 2
                        ws.Cells(self.historyUpdate_rows, integer).Value = lang
                        pathList.append("NULL")
                elif (char_lenght < 0):
                    for i in range(0, abs(char_lenght)):
                        ws.Rows(self.historyUpdate_rows).Delete()
                        del pathList[len(pathList) - 1]

            for idx in range(0, len(pathList)):
            
                jdx = 2
                for value in self.overloading_cell_search_val(sequence= idx, lang= lang):

                    if ("FAIL" in self.overloading_cell_search_val(sequence= idx, lang= lang) or\
                        "N/A" in self.overloading_cell_search_val(sequence= idx, lang= lang) or\
                            "N/T" in self.overloading_cell_search_val(sequence= idx, lang= lang)):
                        ws.Cells(self.historyUpdate_rows, jdx).Value = value
                        jdx= jdx + 1
                self.historyUpdate_rows = self.historyUpdate_rows + 1
            
        else:
            
            getPathList = []
            pathList = []

            for row_idx in range(self.historyUpdate_rows, row):
                print(f"ws.Cells(row_idx, integer).Value : {ws.Cells(row_idx, integer).Value}")
                if (ws.Cells(row_idx, integer).Value == lang):
                    getPathList.append(os.path.basename(ws.Cells(row_idx, 2).Value)) # 파일명만 추출

            for idx in range(0, len(self.create_imgCellCount(lang = lang))):

                data_List = list(self.overloading_cell_search_val(sequence= idx, lang= lang))
                if ("FAIL" in data_List or "N/A" in data_List or "N/T" in data_List):
                    pathList.append(os.path.basename(self.create_imgCellCount(lang)[idx])) # 파일명만 추출모음
                
            print(f'getPathList : {len(getPathList)}')
            print(f'pathList : {len(pathList)}')

            if len(getPathList) != len(pathList):

                print(f'pathList : {len(pathList)}')
                print(f'getPathList : {len(getPathList)}')

                char_lenght = len(pathList) - len(getPathList)
                print(f'char_lenght : {char_lenght}')
                if (char_lenght > 0):
                    
                    for i in range(0, char_lenght):
                        ws.Rows(row).Insert()
                        
                elif (char_lenght < 0):
                    for i in range(0, abs(char_lenght)):
                        ws.Rows(row).Delete()

            row = row - 1
            self.historyUpdate_rows = self.historyUpdate_rows - 1
            print(f'row : {row}')
            for i in range(0, row):
                ws.Rows(self.historyUpdate_rows).Delete()
            
            for i in range(0, row):
                ws.Rows(self.historyUpdate_rows).Insert()

            # 태그적용
            for val in range(0, len(tuples)):
                ws.Cells(self.historyUpdate_rows, val + 1).Value = tuples[val]
                ws.Cells(self.historyUpdate_rows, val + 1).Interior.ColorIndex = self.TABLE_CELL_COLOR # 색상적용
                self.set_Win32com_cellStyle(ws = ws, terget = self.historyUpdate_rows, cell_idx = val + 1, heightSize = None)

                print(f'tuples[val] = {tuples[val]}')
                if tuples[val] == f"{lang}_이미지":
                    print("이미지 사이즈 변경")
                    ws.Range(f"A{val + 1}").Columns.Autofit
                elif tuples[val] in self.evaluation_len(key="Test_List") or tuples[val] == "언어":
                    ws.Columns(val + 1).ColumnWidth = self.SHEET_WIDTH_SHORTSIZE  # 선택 영역 행 크기 설정
                


            # 엑셀 데이터 적용 NA NT FAIL 분류하기
            for key_terget in range(0, len(self.create_imgCellCount(lang=lang))):

                data_List = list(self.overloading_cell_search_val(sequence= key_terget, lang= lang))
                if ("FAIL" in data_List or "N/A" in data_List or "N/T" in data_List):
                    self.historyUpdate_rows = self.historyUpdate_rows + 1
                    resultList = []
                    resultList.append(lang)
                    for key in data_List:
                        resultList.append(key)

                    for val in range(0, len(resultList)):
                        ws.Cells(self.historyUpdate_rows, val + 1).Value = ("" if "" == resultList[val] else resultList[val])
                        self.set_Win32com_cellStyle(ws = ws, terget = self.historyUpdate_rows, cell_idx = val + 1, heightSize = self.IMG_SHEET_HEIGHTSIZE)
                    
                print(f'RESULT 진행도 : {key_terget}/{len(self.create_imgCellCount(lang=lang))}')
                time.sleep(0.005)
                QApplication.processEvents()

            count = 1
            while(True):

                self.historyUpdate_rows = self.historyUpdate_rows + 1        
                print(f"확인 : {ws.Cells(self.historyUpdate_rows, integer).Value}")
                if ws.Cells(self.historyUpdate_rows, integer).Value == lang or ws.Cells(self.historyUpdate_rows, integer).Value == "언어":
                    break
                elif count == 10:
                    print("카운트 멈춤")
                    break
                elif ws.Cells(self.historyUpdate_rows, integer).Value == None:
                    print(f"카운트 적립 : {count}")
                    count = count + 1
                else:
                    ws.Rows(self.historyUpdate_rows).Delete()

            ws.Rows(self.historyUpdate_rows).Insert()
        self.historyUpdate_rows = self.historyUpdate_rows + 1       


                
                
                

